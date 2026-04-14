#!/usr/bin/env python3
"""Import Xiaohongshu note stats from Excel into markdown assets."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from statistics import median
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_STATS_PATH = ROOT / "04-内容数据统计" / "数据统计表.md"
DEFAULT_REPORTS_ROOT = ROOT / "docs" / "reports"
XHS_SECTION_TITLE = "## 小红书数据"


@dataclass(frozen=True)
class XhsNoteRow:
    date: str
    title: str
    genre: str
    views: int
    likes: int
    saves: int
    comments: int
    followers: int
    exposure: int
    ctr: float
    shares: int
    avg_watch_seconds: int

    @property
    def key(self) -> tuple[str, str]:
        return (self.date, self.title)

    @property
    def remark(self) -> str:
        return (
            f"曝光 {self.exposure} / CTR {self.ctr:.1%} / "
            f"分享 {self.shares} / 人均观看时长 {self.avg_watch_seconds}s"
        )

    @property
    def interactions(self) -> int:
        return self.likes + self.saves + self.comments + self.shares

    @property
    def interaction_rate(self) -> float:
        return self.interactions / self.views if self.views else 0.0

    def to_markdown_row(self) -> str:
        return (
            f"| {self.date} | {self.title} | {self.genre} | {self.views} | "
            f"{self.likes} | {self.saves} | {self.comments} | {self.followers} | "
            f"{self.remark} |"
        )


@dataclass(frozen=True)
class ImportResult:
    imported_rows: int
    stats_path: Path
    review_path: Path


def parse_xhs_workbook(xlsx_path: Path) -> list[XhsNoteRow]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Workbook does not contain the expected header rows.")

    headers = rows[1]
    parsed_rows: list[XhsNoteRow] = []
    for raw_row in rows[2:]:
        if not any(value is not None for value in raw_row):
            continue
        item = dict(zip(headers, raw_row))
        published_at = datetime.strptime(item["首次发布时间"], "%Y年%m月%d日%H时%M分%S秒")
        parsed_rows.append(
            XhsNoteRow(
                date=published_at.strftime("%Y-%m-%d"),
                title=str(item["笔记标题"]),
                genre=str(item["体裁"]),
                views=int(item["观看量"] or 0),
                likes=int(item["点赞"] or 0),
                saves=int(item["收藏"] or 0),
                comments=int(item["评论"] or 0),
                followers=int(item["涨粉"] or 0),
                exposure=int(item["曝光"] or 0),
                ctr=float(item["封面点击率"] or 0.0),
                shares=int(item["分享"] or 0),
                avg_watch_seconds=int(item["人均观看时长"] or 0),
            )
        )
    parsed_rows.sort(key=lambda row: (row.date, row.title), reverse=True)
    return parsed_rows


def parse_existing_xhs_rows(markdown: str) -> dict[tuple[str, str], XhsNoteRow]:
    lines = markdown.splitlines()
    start_index = _find_section_start(lines, XHS_SECTION_TITLE)
    if start_index is None:
        return {}
    data_start = start_index + 4
    existing: dict[tuple[str, str], XhsNoteRow] = {}
    for line in lines[data_start:]:
        stripped = line.strip()
        if stripped.startswith("## "):
            break
        if not stripped.startswith("|"):
            continue
        columns = [column.strip() for column in stripped.strip("|").split("|")]
        if len(columns) != 9:
            continue
        if not columns[0]:
            continue
        remark = columns[8]
        exposure, ctr, shares, avg_watch_seconds = _parse_remark(remark)
        row = XhsNoteRow(
            date=columns[0],
            title=columns[1],
            genre=columns[2],
            views=int(columns[3]),
            likes=int(columns[4]),
            saves=int(columns[5]),
            comments=int(columns[6]),
            followers=int(columns[7]),
            exposure=exposure,
            ctr=ctr,
            shares=shares,
            avg_watch_seconds=avg_watch_seconds,
        )
        existing[row.key] = row
    return existing


def merge_xhs_rows(existing_markdown: str, imported_rows: Iterable[XhsNoteRow]) -> list[XhsNoteRow]:
    merged = parse_existing_xhs_rows(existing_markdown)
    for row in imported_rows:
        merged[row.key] = row
    rows = list(merged.values())
    rows.sort(key=lambda row: (row.date, row.title), reverse=True)
    return rows


def update_xhs_section(markdown: str, rows: list[XhsNoteRow]) -> str:
    lines = markdown.splitlines()
    start_index = _find_section_start(lines, XHS_SECTION_TITLE)
    if start_index is None:
        raise ValueError("Could not find `## 小红书数据` section in stats markdown.")

    end_index = len(lines)
    for index in range(start_index + 1, len(lines)):
        if lines[index].startswith("## "):
            end_index = index
            break

    section_lines = [
        XHS_SECTION_TITLE,
        "",
        "| 日期 | 标题 | 类型 | 播放/阅读 | 点赞 | 收藏 | 评论 | 涨粉 | 备注 |",
        "|------|------|------|----------|------|------|------|------|------|",
    ]
    if rows:
        section_lines.extend(row.to_markdown_row() for row in rows)
    else:
        section_lines.append("| | | | | | | | | |")

    updated_lines = lines[:start_index] + section_lines + lines[end_index:]
    return "\n".join(updated_lines).rstrip() + "\n"


def build_review_template(rows: list[XhsNoteRow], *, xlsx_path: Path, stats_path: Path) -> str:
    total_exposure = sum(row.exposure for row in rows)
    total_views = sum(row.views for row in rows)
    total_likes = sum(row.likes for row in rows)
    total_saves = sum(row.saves for row in rows)
    total_comments = sum(row.comments for row in rows)
    total_followers = sum(row.followers for row in rows)
    average_ctr = sum(row.ctr for row in rows) / len(rows) if rows else 0.0
    median_ctr = median([row.ctr for row in rows]) if rows else 0.0
    top_ctr_row = max(rows, key=lambda row: row.ctr, default=None)
    top_save_row = max(rows, key=lambda row: row.saves, default=None)
    low_ctr_rows = sorted(rows, key=lambda row: row.ctr)[:2]

    bottleneck = "点击"
    if rows and average_ctr >= 0.10 and all(row.interaction_rate < 0.08 for row in rows):
        bottleneck = "互动承接"

    lines = [
        f"# 小红书数据复盘模板 - {datetime.now().strftime('%Y-%m-%d')}",
        "",
        "## 来源",
        "",
        f"- Excel：`{xlsx_path}`",
        f"- 数据统计表：`{stats_path}`",
        "",
        "## 数据概览",
        "",
        f"- 导入笔记数：`{len(rows)}`",
        f"- 总曝光 `{total_exposure}`",
        f"- 总观看量 `{total_views}`",
        f"- 总点赞 `{total_likes}`",
        f"- 总收藏 `{total_saves}`",
        f"- 总评论 `{total_comments}`",
        f"- 总涨粉 `{total_followers}`",
        f"- 平均 CTR `{average_ctr:.1%}`",
        f"- 中位 CTR `{median_ctr:.1%}`",
        "",
        "## 主瓶颈初判",
        "",
        f"- 当前主瓶颈：`{bottleneck}`",
        "- 先判断标题、封面、正文前两屏是否承诺一致，再判断是否是平台分发问题。",
        "",
        "## 这轮最值得看",
        "",
    ]
    if top_ctr_row is not None:
        lines.append(
            f"- CTR 最高：`{top_ctr_row.title}` — `{top_ctr_row.ctr:.1%}` / "
            f"`{top_ctr_row.exposure}` 曝光 / `{top_ctr_row.views}` 观看"
        )
    if top_save_row is not None:
        lines.append(
            f"- 收藏最高：`{top_save_row.title}` — `{top_save_row.saves}` 收藏 / "
            f"`{top_save_row.views}` 观看"
        )
    lines.extend(
        [
            "",
            "## 这轮偏弱样本",
            "",
        ]
    )
    for row in low_ctr_rows:
        lines.append(
            f"- `{row.title}` — CTR `{row.ctr:.1%}` / `{row.exposure}` 曝光 / `{row.views}` 观看"
        )
    lines.extend(
        [
            "",
            "## 可直接回写 wiki 的结论",
            "",
            "- 小红书第一钩子优先写“具体问题 + 明确结果”，不要先写抽象概念。",
            "- 标题、封面、正文前两屏必须承诺同一件事。",
            "- 如果数据差，先看点击承诺是否一致，再看正文承接。",
            "",
            "## 下一步建议",
            "",
            "- 选 1 到 2 条 CTR 最弱的笔记，重写标题和封面承诺。",
            "- 把这轮结论回写到 `wiki/方法论/标题创作.md` 与 `wiki/方法论/爆款规律.md`。",
            "- 如果这类 Excel 会持续导出，继续沿用这份模板做每轮复盘。",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def import_xhs_note_stats(*, xlsx_path: Path, stats_path: Path, review_path: Path) -> ImportResult:
    imported_rows = parse_xhs_workbook(xlsx_path)
    stats_markdown = stats_path.read_text(encoding="utf-8")
    merged_rows = merge_xhs_rows(stats_markdown, imported_rows)
    stats_path.write_text(update_xhs_section(stats_markdown, merged_rows), encoding="utf-8")

    review_path.parent.mkdir(parents=True, exist_ok=True)
    review_path.write_text(
        build_review_template(merged_rows, xlsx_path=xlsx_path, stats_path=stats_path),
        encoding="utf-8",
    )
    return ImportResult(
        imported_rows=len(imported_rows),
        stats_path=stats_path,
        review_path=review_path,
    )


def _find_section_start(lines: list[str], title: str) -> int | None:
    for index, line in enumerate(lines):
        if line.strip() == title:
            return index
    return None


def _parse_remark(remark: str) -> tuple[int, float, int, int]:
    match = (
        remark.replace(" ", "")
        .replace("CTR", "CTR")
    )
    exposure_match = _capture_number(match, "曝光")
    ctr_match = _capture_percentage(match, "CTR")
    shares_match = _capture_number(match, "分享")
    watch_match = _capture_number(match, "人均观看时长")
    return exposure_match, ctr_match, shares_match, watch_match


def _capture_number(text: str, label: str) -> int:
    import re

    match = re.search(fr"{label}(\d+)", text)
    return int(match.group(1)) if match else 0


def _capture_percentage(text: str, label: str) -> float:
    import re

    match = re.search(fr"{label}(\d+(?:\.\d+)?)%", text)
    return float(match.group(1)) / 100 if match else 0.0


def default_review_path() -> Path:
    date_str = datetime.now().strftime("%Y-%m-%d")
    return DEFAULT_REPORTS_ROOT / f"xhs-note-review-template-{date_str}.md"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import Xiaohongshu note stats from Excel.")
    parser.add_argument("--xlsx", required=True, help="Path to exported Xiaohongshu Excel file.")
    parser.add_argument(
        "--stats-path",
        default=str(DEFAULT_STATS_PATH),
        help="Target markdown file for 04-内容数据统计/数据统计表.md.",
    )
    parser.add_argument(
        "--review-path",
        default=str(default_review_path()),
        help="Path for generated wiki review template markdown.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = import_xhs_note_stats(
        xlsx_path=Path(args.xlsx),
        stats_path=Path(args.stats_path),
        review_path=Path(args.review_path),
    )
    print(f"Imported {result.imported_rows} rows into {result.stats_path}")
    print(f"Review template written to {result.review_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
